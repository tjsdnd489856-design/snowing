import os
import requests
import re
import sys
from typing import Dict, Optional, Any, List
from concurrent.futures import ThreadPoolExecutor
from dotenv import load_dotenv

# 엑셀 기능을 위한 openpyxl (설치되어 있지 않으면 None 처리)
try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

load_dotenv()

class BaseProvider:
    def fetch(self, gtin: str) -> Optional[Dict[str, Any]]:
        raise NotImplementedError()

class ExcelProvider(BaseProvider):
    """[엑셀 파일 공급자] 로컬 엑셀 파일에서 제품 정보를 검색합니다."""
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.data_cache = {} # 성능을 위해 한 번 읽은 데이터는 메모리에 캐싱
        self._load_excel()

    def _load_excel(self):
        if not load_workbook:
            sys.stderr.write("openpyxl 라이브러리가 없어 엑셀 기능을 사용할 수 없습니다.\n")
            return

        try:
            # data_only=True: 수식 대신 계산된 값을 읽어옴
            wb = load_workbook(self.file_path, data_only=True)
            ws = wb.active
            
            # 헤더 찾기 (1행 가정)
            header = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
            
            # 헤더 매핑 (대소문자 무시, 한글 지원)
            col_map = {h.upper(): i for i, h in enumerate(header) if h}
            
            # 필요한 열 인덱스 찾기
            idx_gtin = col_map.get('GTIN') or col_map.get('바코드')
            idx_name = col_map.get('NAME') or col_map.get('품명') or col_map.get('제품명')
            # '규격', '재고'는 현재 사용하지 않음
            
            if idx_gtin is None:
                sys.stderr.write(f"엑셀 파일({self.file_path})에 '바코드' 또는 'GTIN' 열이 없습니다.\n")
                return

            # 데이터 로드
            count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                # 바코드 읽기 (문자열로 변환)
                gtin_raw = row[idx_gtin]
                if not gtin_raw: continue
                
                # 엑셀에 있는 바코드를 문자열로 변환 (공백 제거)
                # 정수형일 경우 str()로 변환됨 (예: 880123... -> "880123...")
                gtin = str(gtin_raw).strip()
                
                # 품명 읽기
                name_raw = row[idx_name] if idx_name is not None else ""
                full_name = str(name_raw).strip() if name_raw else "Unknown Product"
                
                # 품명에서 도수 분리
                name, power = self._parse_name_and_power(full_name)
                
                self.data_cache[gtin] = {
                    "name": name,
                    "power": power
                }
                count += 1
            
            print(f"[DEBUG] 엑셀 로드 완료: {count}개의 데이터가 캐시되었습니다.")
                
        except Exception as e:
            sys.stderr.write(f"엑셀 파일 로드 실패: {self.file_path} ({e})\n")

    def _parse_name_and_power(self, full_name: str) -> tuple[str, str]:
        """
        품명 문자열에서 도수를 추출합니다.
        예: "바이오피니티 -3.00" -> ("바이오피니티", "-3.00")
        예: "아큐브 오아시스 +1.25" -> ("아큐브 오아시스", "+1.25")
        """
        # 도수 패턴: + 또는 - 부호가 있거나 없으며, 숫자.숫자 형식 (예: -3.00, +1.25, 0.00)
        # 또는 단순 정수형 도수 (예: -3, +2)
        power_pattern = r'([+-]?\d+\.\d{2}|[+-]?\d+\.\d+|[+-]\d+)'
        
        match = re.search(power_pattern, full_name)
        if match:
            power = match.group(1)
            # 도수를 제외한 나머지 문자열을 제품명으로 사용
            name = full_name.replace(power, "").strip().strip("-_, ")
            return name, power
        
        return full_name, "N/A"

    def fetch(self, gtin: str) -> Optional[Dict[str, Any]]:
        # 1. 있는 그대로 검색 (14자리 또는 13자리)
        print(f"[DEBUG] 엑셀 검색 시도 (원본): {gtin}")
        data = self.data_cache.get(gtin)
        if data:
            print(f"[DEBUG] 엑셀 검색 성공 (원본): {data['name']}")
            return data
            
        # 2. 만약 입력된 바코드가 14자리이고 0으로 시작하면, 13자리로 변환해서 재검색
        # (엑셀에는 13자리로 저장되어 있을 가능성이 높음)
        if len(gtin) == 14 and gtin.startswith('0'):
            gtin_13 = gtin[1:]
            print(f"[DEBUG] 엑셀 재검색 시도 (13자리): {gtin_13}")
            data = self.data_cache.get(gtin_13)
            if data:
                print(f"[DEBUG] 엑셀 검색 성공 (13자리): {data['name']}")
                return data
            
        print("[DEBUG] 엑셀 검색 실패: 데이터 없음")
        return None

class MFDSProvider(BaseProvider):
    """[식약처 API 공급자] 병렬 호출을 사용하여 속도를 극대화합니다."""
    
    def __init__(self, api_key: str):
        self.api_key = api_key
        self.base_url = "https://apis.data.go.kr/1471000"

    def fetch(self, gtin: str) -> Optional[Dict[str, Any]]:
        # 두 개의 엔드포인트를 동시에 호출하여 속도 향상
        with ThreadPoolExecutor(max_workers=2) as executor:
            future_base = executor.submit(self._call, "MdeqStdCdUnityInfoService01", "getMdeqStdCdUnityInfoInq01", gtin)
            future_detail = executor.submit(self._call, "MdvUdiInfoService", "getMdvUdiInfoInq01", gtin)
            
            content_base = future_base.result()
            content_detail = future_detail.result()

        final_content = content_detail if content_detail else content_base
        if not final_content:
            return None

        return self._extract(final_content)

    def _call(self, service: str, endpoint: str, gtin: str) -> Optional[str]:
        url = f"{self.base_url}/{service}/{endpoint}"
        params = {'serviceKey': self.api_key, 'type': 'json', 'pageNo': 1, 'numOfRows': 1, 'UDIDI_CD': gtin}
        try:
            resp = requests.get(url, params=params, timeout=5) # 타임아웃 단축
            if resp.status_code == 200 and '"totalCount":0' not in resp.text:
                return resp.text
        except: pass
        return None

    def _extract(self, content: str) -> Optional[Dict[str, Any]]:
        fields = ["PRDT_ADD_EXPL", "PRDT_NM", "MODEL_NM"]
        for f in fields:
            m = re.search(rf'{f}["\>\]\s:]+([^"<\n]+)', content, re.IGNORECASE)
            if m:
                raw = m.group(1).strip().strip('"} ,;')
                p_m = re.search(r'([+-]?\d+\.\d{2})', raw)
                power = p_m.group(1) if p_m else "N/A"
                name = raw.replace(power, "").strip("- ").strip()
                return {"name": name or "식약처 등록 제품", "power": power}
        return None

class APIClient:
    def __init__(self):
        self.providers: List[BaseProvider] = []
        
        # 1. 엑셀 파일 우선 검색 (LENS_EXCEL_PATH 환경변수 사용)
        # 기본값으로 'product_list.xlsx'도 확인하도록 수정
        excel_path = os.getenv("LENS_EXCEL_PATH", "product_list.xlsx").strip()
        if excel_path and os.path.exists(excel_path):
            print(f"[DEBUG] 엑셀 파일 로드 시도: {excel_path}")
            self.providers.append(ExcelProvider(excel_path))
        else:
            print(f"[DEBUG] 엑셀 파일을 찾을 수 없음: {excel_path}")
        
        # 2. 식약처 API 검색
        mfds_key = os.getenv("LENS_API_KEY", "").strip()
        if mfds_key:
            self.providers.append(MFDSProvider(mfds_key))

    def fetch_product_info(self, gtin: str) -> Optional[Dict[str, Any]]:
        if not gtin: return None
        for provider in self.providers:
            info = provider.fetch(gtin)
            if info:
                return {**info, 'gtin': gtin, 'source': provider.__class__.__name__}
        return None

    def sync_with_local_db(self, api_data: Dict, local_data: Dict) -> Dict:
        synced = local_data.copy()
        if api_data:
            synced.update({'name': api_data.get('name'), 'power': api_data.get('power')})
        return synced
